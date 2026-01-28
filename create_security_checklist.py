from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_security_checklist(filename="security_checklist.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Security Checklist"

    # Headers
    headers = ["Category", "ID", "Checklist Item", "Status", "Priority", "Notes"]
    ws.append(headers)

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD") # Blue
    alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                   top=Side(style='thin'), bottom=Side(style='thin'))

    # Apply Header Styles
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Data (OWASP Top 10 & Best Practices)
    checklist_data = [
        ("A01: Broken Access Control", "BAC-001", "Ensure least privilege principle is applied.", "Pending", "High", ""),
        ("A01: Broken Access Control", "BAC-002", "Verify IDOR protections on all endpoints.", "Pending", "High", ""),
        ("A02: Cryptographic Failures", "CRY-001", "Ensure no hardcoded secrets/keys in code.", "Pending", "Critical", ""),
        ("A02: Cryptographic Failures", "CRY-002", "Use strong encryption for data at rest.", "Pending", "High", ""),
        ("A03: Injection", "INJ-001", "Use parameterized queries (Prepared Statements).", "Pending", "Critical", ""),
        ("A03: Injection", "INJ-002", "Sanitize all user inputs.", "Pending", "High", ""),
        ("A04: Insecure Design", "DES-001", "Perform Threat Modeling.", "Pending", "Medium", ""),
        ("A05: Security Misconfiguration", "MIS-001", "Disable debug mode in production.", "Pending", "High", ""),
        ("A06: Vuln & Outdated Components", "OLD-001", "Check dependencies for known vulnerabilities.", "Pending", "High", ""),
        ("A07: Identification & Auth Fail", "AUTH-001", "Implement Multi-Factor Authentication (MFA).", "Pending", "High", ""),
        ("A08: Software & Data Integrity", "INT-001", "Verify CI/CD pipeline integrity.", "Pending", "Medium", ""),
        ("A09: Logging & Monitoring", "LOG-001", "Ensure login failures are logged.", "Pending", "Medium", ""),
        ("A10: SSRF", "SSRF-001", "Validate all URLs in inputs.", "Pending", "High", ""),
        ("General", "ENV-001", "Use Environment Variables for config.", "Pending", "High", ""),
    ]

    # Write Data
    for row_data in checklist_data:
        ws.append(row_data)

    # Styling Content
    for row in ws.iter_rows(min_row=2, max_row=len(checklist_data)+1):
        for cell in row:
            cell.alignment = alignment
            cell.border = border

    # Column Widths
    column_widths = [25, 10, 50, 10, 10, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # Add AutoFilter
    ws.auto_filter.ref = ws.dimensions

    wb.save(filename)
    print(f"✅ Security Checklist created: {filename}")

if __name__ == "__main__":
    create_security_checklist()
